import streamlit as st
import pandas as pd
import psycopg2
import plotly.graph_objects as go
import datetime as dt 
from dateutil.relativedelta import relativedelta
from streamlit_plotly_events import plotly_events
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl import load_workbook
from io import BytesIO
from itertools import product

# ------------------- Page Setup -------------------
st.set_page_config(page_title="Planned vs Actual", layout="wide")
st.title("Anudesh Annotation Dashboard")
st.markdown("---")

# ------------------- Report Index -------------------
st.markdown("""
### 📊 Report Index  
🔹 [Chart 1: Planned vs Actual Annotations (by Category & Script Type)](#chart-1)  
🔹 [Chart 2: Planned vs Actual Annotations (by Language & Script Type)](#chart-2)  
🔹 [Chart 3: Daily User Progress ](#chart-3)  
🔹 [Chart 4: Task Status by Target Language](#chart-4)  
🔹 [Chart 5: Workspace level Annotation Stats](#chart-5)
""", unsafe_allow_html=True)

# ------------------- DB Config -------------------
# DB_HOST = 'localhost'
# DB_PORT = "5432"
# DB_NAME = 'Summary_Report'
# DB_USER = 'postgres'
# DB_PASS = 'postgres'

# @st.cache_resource
# def get_connection():
#     try:
#         conn = psycopg2.connect(
#             host=DB_HOST,
#             port=DB_PORT,
#             dbname=DB_NAME,
#             user=DB_USER,
#             password=DB_PASS
#         )
#         st.success(" Successfully connected to the database.")
#         return conn
#     except Exception as e:
#         st.error(f" Database connection failed: {e}")
#         return None
# conn = get_connection()


# @st.cache_resource
# def get_connection():
#     try:
#         conn = psycopg2.connect(
#             host=st.secrets["DB_HOST"],
#             port=st.secrets["DB_PORT"],
#             dbname=st.secrets["DB_NAME"],
#             user=st.secrets["DB_USER"],
#             password=st.secrets["DB_PASS"]
#         )
#         st.success("✅ Successfully connected to the database.")
#         return conn
#     except Exception as e:
#         st.error(f"❌ Database connection failed: {e}")
#         st.stop()

# conn = get_connection()

DB_HOST = "turntable.proxy.rlwy.net"
DB_PORT = "17356"
DB_NAME = "railway"
DB_USER = "postgres"
DB_PASSWORD = "EvFmtpzsUtMeblVMYtZgOSrDnVlehUVU"
import streamlit as st
import psycopg2

conn = psycopg2.connect(
    host=st.secrets["database"]["host"],
    port=st.secrets["database"]["port"],
    dbname=st.secrets["database"]["name"],
    user=st.secrets["database"]["user"],
    password=st.secrets["database"]["password"]
)


st.markdown("---")

# ------------------- Time Filter Helper -------------------
def apply_time_filter(df, time_filter):
    today = dt.date.today()
    start_date, end_date = None, today

    if time_filter == "Today":
        start_date = today
    elif time_filter == "Yesterday":
        start_date = today - dt.timedelta(days=1)
        end_date = start_date
    elif time_filter == "Last Week":
        start_date = today - dt.timedelta(days=today.weekday() + 7)
        end_date = start_date + dt.timedelta(days=6)
    elif time_filter == "Last Month":
        first_day_this_month = today.replace(day=1)
        start_date = (first_day_this_month - dt.timedelta(days=1)).replace(day=1)
        end_date = first_day_this_month - dt.timedelta(days=1)
    elif time_filter == "Custom Range":
        col3, col4 = st.columns(2)
        with col3:
            start_date = st.date_input("Start Date", value=today - dt.timedelta(days=30), key="s1")
        with col4:
            end_date = st.date_input("End Date", value=today, key="e1")

    if time_filter != "All Time" and start_date:
        df = df[(df["updated_at"].dt.date >= start_date) & (df["updated_at"].dt.date <= end_date)]
    return df

df_debug = pd.read_sql("SELECT * FROM intermediate_table LIMIT 5", conn)
st.write("Intermediate table sample:")
st.write(df_debug)

# ------------------- Chart 1: By Category & Script Type -------------------
# query_script_type = """
# SELECT 
#     category,
#     script_type,
#     SUM(planned_no_of_annotations) AS planned,
#     SUM(actual_no_of_annotations) AS actual,
#     MIN(updated_at) AS updated_at
# FROM intermediate_table
# GROUP BY category, script_type
# ORDER BY category, script_type;
# """
# query_script_type = """SELECT 
#     category,
#     script_type,
#     SUM(planned) AS planned,
#     SUM(actual) AS actual
# FROM language_table
# WHERE language IS NOT NULL 
#   AND TRIM(language) <> ''
#   AND LOWER(language) NOT IN ('filipino', 'english')
# GROUP BY category, script_type
# ORDER BY category, script_type;
# """

query_script_type = """
SELECT 
    category,
    script_type,
    SUM(planned) AS planned,
    SUM(actual) AS actual,
    MIN(updated_at) AS updated_at
FROM language_table
GROUP BY category, script_type
ORDER BY category, script_type;
"""

df_script_type_all = pd.read_sql(query_script_type, conn)
print(df_script_type_all.columns)

df_script_type_all["updated_at"] = pd.to_datetime(df_script_type_all["updated_at"], errors='coerce')


df_script_type_all["updated_at"] = pd.to_datetime(df_script_type_all["updated_at"])

planned_reference_df = df_script_type_all[["category", "script_type", "planned"]].copy()
all_categories = df_script_type_all["category"].unique()
all_scripts = df_script_type_all["script_type"].unique()
full_index = pd.MultiIndex.from_product([all_categories, all_scripts], names=["category", "script_type"])
planned_reference_df = planned_reference_df.set_index(["category", "script_type"]).reindex(full_index, fill_value=0).reset_index()

col1, col2 = st.columns([2, 1])
with col1:
    time_filter1 = st.radio(
        "⏱ Time Range",
        ["All Time", "Today", "Yesterday", "Last Week", "Last Month", "Custom Range"],
        key="time1",
        horizontal=True
    )
with col2:
    categories1 = sorted(df_script_type_all["category"].unique())
    selected_categories1 = st.multiselect(" Select Categories", options=categories1, default=categories1, key="cat1")

st.markdown("### <a name='chart-1'></a>📈 Chart 1: Planned vs Actual Annotations (by Category & Script Type)", unsafe_allow_html=True)

df1_filtered_actual = apply_time_filter(df_script_type_all.copy(), time_filter1)
df1_filtered_actual = df1_filtered_actual[df1_filtered_actual["category"].isin(selected_categories1)]

df1_filtered = df1_filtered_actual.drop(columns=["planned"]).merge(planned_reference_df, on=["category", "script_type"], how="right")
df1_filtered = df1_filtered[df1_filtered["category"].isin(selected_categories1)].fillna({"actual": 0})

script_types = sorted([s for s in df1_filtered["script_type"].dropna().unique()])
categories = sorted(selected_categories1)
color_map = {
    # Blue shades for English
    ("English", "planned"): "#B3D4FC",   # Light Blue
    ("English", "actual"): "#1F77B4",    # Medium/Darker Blue

    # Green shades for Native
    ("Native", "planned"): "#BDEDB3",    # Light Green
    ("Native", "actual"): "#2CA02C",     # Medium/Darker Green

    # Yellow shades for Indic Roman
    ("Indic Roman", "planned"): "#FFF4B3",  # Light Yellow
    ("Indic Roman", "actual"): "#FFBB00",   # Medium/Darker Yellow
}


fig1 = go.Figure()
for script in script_types:
    for typ in ["planned", "actual"]:
        filtered = df1_filtered[df1_filtered["script_type"] == script]
        y_vals = []
        for cat in categories:
            val = filtered[(filtered["category"] == cat)][typ].values
            val = val[0] if len(val) > 0 else 0
            y_vals.append(val)
        fig1.add_trace(go.Bar(
            x=categories,
            y=y_vals,
            name=f"{script} ({typ.capitalize()})",
            marker_color=color_map.get((script, typ), "#cccccc"),
            
            offsetgroup=typ,
            legendgroup=script,
            text=["0" if v == 0 else f"{v:.0f}" for v in y_vals],
            textposition="outside",
            texttemplate="%{text}",
            cliponaxis=False
        ))

fig1.update_layout(
    xaxis_title="Category",
    yaxis_title="Annotation Count",
    barmode="stack",
    legend_title="Script Type & Annotation Type",
    plot_bgcolor="rgba(0,0,0,0)",
    paper_bgcolor="rgba(0,0,0,0)",
    bargap=0.25
)
st.plotly_chart(fig1, use_container_width=True)

download_df = df1_filtered_actual.copy()  # This contains the filtered actuals based on time/category

csv = download_df.to_csv(index=False).encode('utf-8')

st.download_button(
    label="⬇️ Download Chart 1 data as CSV",
    data=csv,
    file_name="filtered_annotations.csv",
    mime="text/csv"
)



# ------------------- Chart 1 Drilldown -------------------
st.markdown("### 🔍 Category Drilldown (Dropdown)")

category_options = ["-- Select a category --"] + sorted(categories)
selected_category = st.selectbox("Select a Category to Drill Down", options=category_options)

if selected_category != "-- Select a category --":
    detail_df = df1_filtered[df1_filtered["category"] == selected_category]
    detail_fig = go.Figure()
    detail_fig.add_trace(go.Bar(
        x=detail_df["script_type"],
        y=detail_df["planned"],
        name="Planned",
        marker_color="#5DADE2",
        text=detail_df["planned"],
        textposition="auto"
    ))
    detail_fig.add_trace(go.Bar(
        x=detail_df["script_type"],
        y=detail_df["actual"],
        name="Actual",
        marker_color="#48C9B0",
        text=detail_df["actual"],
        textposition="auto"
    ))
    detail_fig.update_layout(
        title=f"Detailed Breakdown for '{selected_category}'",
        xaxis_title="Script Type",
        yaxis_title="Annotation Count",
        barmode="group",
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)"
    )
    st.plotly_chart(detail_fig, use_container_width=True)



# ------------------- Chart 2: By Language & Script Type -------------------
st.markdown("---")
st.markdown("### <a name='chart-2'></a>📈 Chart 2: Planned vs Actual Annotations (by Language & Script Type)", unsafe_allow_html=True)


query_lang = """
SELECT 
    category,
    language,
    script_type,
    SUM(planned) AS planned,
    SUM(actual) AS actual,
    MIN(updated_at) AS updated_at 
FROM language_table
WHERE language IS NOT NULL AND language <> '' AND LOWER(language) NOT IN ('filipino', 'english')
GROUP BY category, language, script_type 
ORDER BY language, script_type;
"""


df_lang = pd.read_sql(query_lang, conn)
import datetime

if df_lang["updated_at"].dtype == "object" or isinstance(df_lang["updated_at"].iloc[0], datetime.time):
    df_lang["updated_at"] = df_lang["updated_at"].apply(
        lambda t: datetime.datetime.combine(datetime.date.today(), t)
        if isinstance(t, datetime.time)
        else pd.to_datetime(t, errors='coerce')  # fallback for normal datetime strings
    )
else:
    df_lang["updated_at"] = pd.to_datetime(df_lang["updated_at"], errors='coerce')

df_lang["updated_at"] = pd.to_datetime(df_lang["updated_at"])

col5, col6 = st.columns([2, 1])
with col5:
    time_filter2 = st.radio(
        "⏱ Time Range",
        ["All Time", "Today", "Yesterday", "Last Week", "Last Month", "Custom Range"],
        key="time2",
        horizontal=True
    )

with col6:
    categories2 = ["All"] + sorted(df_lang["category"].dropna().unique())
    selected_category2 = st.selectbox("Select Category", options=categories2, key="cat2")

    script_types2 = ["All"] + sorted(df_lang["script_type"].dropna().unique())
    selected_script_type2 = st.selectbox("Select Script Type", options=script_types2, key="stype2")

    languages2 = ["All"] + sorted(df_lang["language"].dropna().unique())
    selected_language2 = st.selectbox("Select Language", options=languages2, key="lang2")


df2_filtered = apply_time_filter(df_lang.copy(), time_filter2)

if selected_category2 != "All":
    df2_filtered = df2_filtered[df2_filtered["category"] == selected_category2]

if selected_script_type2 != "All":
    df2_filtered = df2_filtered[df2_filtered["script_type"] == selected_script_type2]

if selected_language2 != "All":
    df2_filtered = df2_filtered[df2_filtered["language"] == selected_language2]

df2_filtered.dropna(subset=["language", "script_type", "planned", "actual"], inplace=True)
df2_filtered["planned"] = pd.to_numeric(df2_filtered["planned"], errors="coerce")
df2_filtered["actual"] = pd.to_numeric(df2_filtered["actual"], errors="coerce")
df2_filtered["language"] = df2_filtered["language"].apply(lambda x: ', '.join(x) if isinstance(x, list) else str(x))

fig3 = go.Figure()
script_types = sorted([s for s in df2_filtered["script_type"].dropna().unique()])
languages = df2_filtered["language"].unique()

for stype in script_types:
    for typ in ["planned", "actual"]:
        subset = df2_filtered[df2_filtered["script_type"] == stype]
        y_vals = []
        langs = []
        for lang in languages:
            value = subset[subset["language"] == lang][typ].sum()
            if value > 0:
                y_vals.append(value)
                langs.append(lang)
        fig3.add_trace(go.Bar(
            x=langs,
            y=y_vals,
            name=f"{typ.capitalize()} - {stype}",
            marker_color=color_map.get((stype, typ), "#cccccc"),
            

            offsetgroup=typ,
            legendgroup=stype,
            text=y_vals,
            textposition="auto"
        ))

fig3.update_layout(
    xaxis_title="Language",
    yaxis_title="Annotation Count",
    barmode="stack",
    plot_bgcolor="rgba(0,0,0,0)",
    paper_bgcolor="rgba(0,0,0,0)",
    legend_title="Script Type + Type"
)
st.plotly_chart(fig3, use_container_width=True)

# -------------------------------------------------------------

# st.download_button(
#     label="📅 Download Chart 2 Data as CSV",
#     data=df2_filtered.to_csv(index=False),
#     file_name="language_script_type_annotations.csv",
#     mime="text/csv"
# )

# ------------------------ Hardcoded Excel Layout ------------------------


# Load the Excel template
template_path = "language_report.xlsx"  # Ensure this file is present in working dir
wb = load_workbook(template_path)
ws = wb.active

# Define category + script_type + value_type -> column number
column_map = {
    ("Creative", "English", "planned"): 2,
    ("Creative", "English", "actual"): 3,
    ("Creative", "Native", "planned"): 4,
    ("Creative", "Native", "actual"): 5,
    ("Creative", "Indic Roman", "planned"): 6,
    ("Creative", "Indic Roman", "actual"): 7,
    ("Cultural", "English", "planned"): 8,
    ("Cultural", "English", "actual"): 9,
    ("Cultural", "Native", "planned"): 10,
    ("Cultural", "Native", "actual"): 11,
    ("Cultural", "Indic Roman", "planned"): 12,
    ("Cultural", "Indic Roman", "actual"): 13,
    ("Education", "English", "planned"): 14,
    ("Education", "English", "actual"): 15,
    ("Education", "Native", "planned"): 16,
    ("Education", "Native", "actual"): 17,
    ("Education", "Indic Roman", "planned"): 18,
    ("Education", "Indic Roman", "actual"): 19,
    ("General", "English", "planned"): 20,
    ("General", "English", "actual"): 21,
    ("General", "Native", "planned"): 22,
    ("General", "Native", "actual"): 23,
    ("General", "Indic Roman", "planned"): 24,
    ("General", "Indic Roman", "actual"): 25,
    ("Safety", "English", "planned"): 26,
    ("Safety", "English", "actual"): 27,
    ("Safety", "Native", "planned"): 28,
    ("Safety", "Native", "actual"): 29,
    ("Safety", "Indic Roman", "planned"): 30,
    ("Safety", "Indic Roman", "actual"): 31,
}

# Define row numbers by language
languages_list = [
    "Assamese", "Bengali", "Bodo", "Dogri", "Gujarati", "Hindi", "Kannada", "Kashmiri",
    "Maithili", "Malayalam", "Marathi", "Nepali", "Odia", "Punjabi", "Santali", "Tamil",
    "Telugu", "Urdu"
]
language_to_row = {lang: idx + 4 for idx, lang in enumerate(languages_list)}  # Excel starts at row 4

# Fill Excel Sheet
# for _, row in df_lang.iterrows():
for _, row in df2_filtered.iterrows():
    lang = str(row["language"]).strip()
    cat = str(row["category"]).strip()
    stype = str(row["script_type"]).strip()
    planned = pd.to_numeric(row["planned"], errors="coerce")
    actual = pd.to_numeric(row["actual"], errors="coerce")

    if lang in language_to_row:
        row_num = language_to_row[lang]
        for val_type, val in [("planned", planned), ("actual", actual)]:
            col_num = column_map.get((cat, stype, val_type))
            if col_num and pd.notna(val):
                ws.cell(row=row_num, column=col_num).value = int(val)

# Green fill style for highlighting
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

# Apply conditional formatting for all relevant rows and columns
for (category, script_type, _), planned_col in column_map.items():
    actual_col = column_map.get((category, script_type, "actual"))
    if not actual_col:
        continue  # skip if actual column not found

    for lang in languages_list:
        row = language_to_row[lang]

        planned_cell = ws.cell(row=row, column=planned_col).coordinate
        actual_cell = ws.cell(row=row, column=actual_col).coordinate

        # Highlight actual cell if actual > planned
        formula = f'{actual_cell}>{planned_cell}'
        ws.conditional_formatting.add(actual_cell, CellIsRule(operator='greaterThan', formula=[planned_cell], fill=green_fill))

# Export Excel in-memory
output = BytesIO()
wb.save(output)
output.seek(0)

# Download Button
st.download_button(
    label="📥 Download High level Language Report ",
    data=output,
    file_name="language_report_filled.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)



#--------------------------------------

# from openpyxl import Workbook
# from openpyxl.utils.dataframe import dataframe_to_rows
# from io import BytesIO

# # Create workbook and summary sheet
# wb = Workbook()
# summary_ws = wb.active
# summary_ws.title = "Summary"

# # Add headers to summary
# summary_headers = ["Language", "Category", "Script Type", "Planned", "Actual", "Last Updated"]
# summary_ws.append(summary_headers)

# # Add rows to summary
# for _, row in df2_filtered.iterrows():
#     summary_ws.append([
#         row["language"],
#         row["category"],
#         row["script_type"],
#         int(row["planned"]) if pd.notna(row["planned"]) else 0,
#         int(row["actual"]) if pd.notna(row["actual"]) else 0,
#         row["updated_at"].strftime("%Y-%m-%d") if pd.notna(row["updated_at"]) else ""
#     ])

# # Create individual sheets per language
# for lang in df2_filtered["language"].unique():
#     lang_df = df2_filtered[df2_filtered["language"] == lang].copy()
#     lang_df = lang_df[["category", "script_type", "planned", "actual", "updated_at"]]
#     lang_df["planned"] = lang_df["planned"].fillna(0).astype(int)
#     lang_df["actual"] = lang_df["actual"].fillna(0).astype(int)
#     lang_df["updated_at"] = lang_df["updated_at"].dt.strftime("%Y-%m-%d")

#     ws = wb.create_sheet(title=lang[:31])  # Excel sheet names max length = 31
#     ws.append(["Category", "Script Type", "Planned", "Actual", "Last Updated"])

#     for r in dataframe_to_rows(lang_df, index=False, header=False):
#         ws.append(r)

# # Save to memory
# excel_bytes = BytesIO()
# wb.save(excel_bytes)
# excel_bytes.seek(0)

# # Download button
# st.download_button(
#     label="📥 Download Multi-Sheet Excel Report",
#     data=excel_bytes,
#     file_name="language_wise_report.xlsx",
#     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# )

# ------------------- Language Drilldown Dropdown -------------------
st.markdown("### 🔍 Language Drilldown (Dropdown)")
# selected_language = st.selectbox("Select a Language to Drill Down", options=sorted(languages))
available_langs = sorted(df2_filtered["language"].unique())
selected_language = st.selectbox("Select a Language to Drill Down", options=available_langs)


if selected_language:
    lang_df = df2_filtered[df2_filtered["language"] == selected_language].copy()
    lang_df = lang_df.groupby("script_type")[["planned", "actual"]].sum().reset_index()

    # Sort by planned value for consistent bar alignment
    lang_df.sort_values("planned", ascending=False, inplace=True)

    lang_fig = go.Figure()
    lang_fig.add_trace(go.Bar(
        x=lang_df["script_type"],
        y=lang_df["planned"],
        name="Planned",
        marker_color="#5DADE2",
        text=lang_df["planned"],
        textposition="auto"
    ))
    lang_fig.add_trace(go.Bar(
        x=lang_df["script_type"],
        y=lang_df["actual"],
        name="Actual",
        marker_color="#48C9B0",
        text=lang_df["actual"],
        textposition="auto"
    ))
    lang_fig.update_layout(
        title=f"📌 Detailed Breakdown for Language: '{selected_language}'",
        xaxis_title="Script Type",
        yaxis_title="Annotation Count",
        barmode="group",  # <-- Grouped bars instead of stacked
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        legend_title="Annotation Type"
    )
    st.plotly_chart(lang_fig, use_container_width=True)


# ------------------- Chart 3: Daily User Progress (Filtered) -------------------
import streamlit as st
import pandas as pd
import datetime as dt
import plotly.graph_objects as go

st.markdown("---")
st.markdown("### <a name='chart-3'></a>📈 Chart 3: Daily User Progress (Filtered)", unsafe_allow_html=True)

# --- Initial Full Data Fetch for All Dropdowns ---
base_query = "SELECT DISTINCT language, user_first_name, category, CAST(workspace_id AS TEXT) AS workspace_id FROM intermediate_table WHERE language IS NOT NULL AND user_first_name IS NOT NULL AND category IS NOT NULL AND workspace_id IS NOT NULL"
df_filters = pd.read_sql(base_query, conn)

# --- Sidebar Filters Logic (Cascading Dropdowns) ---
st.markdown("#### 🔎 Apply Filters")

col1, col2, col3, col4 = st.columns(4)

with col1:
    selected_language = st.selectbox("🗣️ Select Language", ["All"] + sorted(df_filters["language"].unique().tolist()), index=0, key="c3_language")

# Filter dataframe after language selection
df_filtered = df_filters.copy()
if selected_language != "All":
    df_filtered = df_filtered[df_filtered["language"] == selected_language]

with col2:
    users = sorted(df_filtered["user_first_name"].unique().tolist())
    selected_user = st.selectbox("👤 Select User", ["All"] + users, index=0, key="c3_user")

if selected_user != "All":
    df_filtered = df_filtered[df_filtered["user_first_name"] == selected_user]

with col3:
    categories = sorted(df_filtered["category"].unique().tolist())
    selected_category = st.selectbox("🏷️ Select Category", ["All"] + categories, index=0, key="c3_category")

if selected_category != "All":
    df_filtered = df_filtered[df_filtered["category"] == selected_category]

with col4:
    workspaces = sorted(df_filtered["workspace_id"].unique().tolist())
    selected_workspace = st.selectbox("🗂️ Select Workspace ID", ["All"] + workspaces, index=0, key="c3_workspace")

# --- Time Filter ---
time_filter = st.radio("⏱ Time Range", ["All Time", "Today", "Yesterday", "Last Week", "Last Month", "Custom Range"],
                       horizontal=True, key="c3_time")

# --- Time Logic ---
today = dt.date.today()
start_date, end_date = None, today

if time_filter == "Today":
    start_date = today
elif time_filter == "Yesterday":
    start_date = today - dt.timedelta(days=1)
    end_date = start_date
elif time_filter == "Last Week":
    start_date = today - dt.timedelta(days=today.weekday() + 7)
    end_date = start_date + dt.timedelta(days=6)
elif time_filter == "Last Month":
    first_day_this_month = today.replace(day=1)
    start_date = (first_day_this_month - dt.timedelta(days=1)).replace(day=1)
    end_date = first_day_this_month - dt.timedelta(days=1)
elif time_filter == "Custom Range":
    col5, col6 = st.columns(2)
    with col5:
        start_date = st.date_input("Start Date", value=today - dt.timedelta(days=30), key="c3_start")
    with col6:
        end_date = st.date_input("End Date", value=today, key="c3_end")

# --- Build WHERE Clause ---
where_clauses = []
params = []

if selected_language != "All":
    where_clauses.append("language = %s")
    params.append(selected_language)

if selected_user != "All":
    where_clauses.append("user_first_name = %s")
    params.append(selected_user)

if selected_category != "All":
    where_clauses.append("category = %s")
    params.append(selected_category)

if selected_workspace != "All":
    where_clauses.append("CAST(workspace_id AS TEXT) = %s")
    params.append(selected_workspace)

if time_filter != "All Time" and start_date:
    where_clauses.append("DATE(updated_at) BETWEEN %s AND %s")
    params.extend([start_date, end_date])

where_clause = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

# --- Final SQL Query ---
query_progress = f"""
SELECT 
    DATE(updated_at) AS work_date, 
    user_first_name,
    category,
    COUNT(*) AS annotations_done
FROM intermediate_table
{where_clause}
GROUP BY work_date, user_first_name, category
ORDER BY work_date ASC;
"""

# --- Load & Process Data ---
df_progress = pd.read_sql(query_progress, conn, params=params)
df_progress["work_date"] = pd.to_datetime(df_progress["work_date"])

# --- Plotting ---
fig4 = go.Figure()
fig4.add_trace(go.Scatter(
    x=df_progress["work_date"],
    y=df_progress["annotations_done"],
    mode="lines+markers",
    name="Annotations Done",
    marker=dict(color="#F39C12"),
    line=dict(width=2)
))

# --- Dynamic Title ---
title_text = "📅 Daily User Progress"
if selected_user != "All":
    title_text += f" for {selected_user}"
if selected_category != "All":
    title_text += f" in '{selected_category}'"
if selected_workspace != "All":
    title_text += f" (Workspace ID: {selected_workspace})"
if selected_language != "All":
    title_text += f" in language '{selected_language}'"
if time_filter != "All Time":
    title_text += f" ({start_date} to {end_date})"

fig4.update_layout(
    title=title_text,
    xaxis_title="Date",
    yaxis_title="Annotations Done",
    plot_bgcolor="rgba(0,0,0,0)",
    paper_bgcolor="rgba(0,0,0,0)"
)

# --- Display Chart ---
st.plotly_chart(fig4, use_container_width=True, key="chart3_fig")

# --- Download Button ---
st.download_button(
    label="⬇️ Download Chart 3 Data as CSV",
    data=df_progress.to_csv(index=False),
    file_name="daily_user_progress_filtered.csv",
    mime="text/csv",
    key="chart3_download"
)


# # ------------------- Chart 3: Daily User Progress (Filtered) -------------------
# st.markdown("---")
# st.markdown("### <a name='chart-3'></a>📈 Chart 3: Daily User Progress (Filtered)", unsafe_allow_html=True)

# # --- Fetch Distinct Filter Options ---
# usernames = ["All"] + pd.read_sql(
#     "SELECT DISTINCT user_first_name FROM intermediate_table WHERE user_first_name IS NOT NULL ORDER BY user_first_name",
#     conn)["user_first_name"].tolist()

# categories = ["All"] + pd.read_sql(
#     "SELECT DISTINCT category FROM intermediate_table WHERE category IS NOT NULL ORDER BY category",
#     conn)["category"].tolist()

# workspace_ids = ["All"] + pd.read_sql(
#     "SELECT DISTINCT workspace_id FROM intermediate_table WHERE workspace_id IS NOT NULL ORDER BY workspace_id",
#     conn)["workspace_id"].astype(str).tolist()

# languages = ["All"] + pd.read_sql(
#     "SELECT DISTINCT language FROM intermediate_table WHERE language IS NOT NULL ORDER BY language",
#     conn)["language"].tolist()


# # --- Sidebar Filters ---
# col1, col2, col3 = st.columns([2, 2, 3])

# with col1:
#     selected_user = st.selectbox("👤 Select Username", usernames, index=0, key="chart3_user")

# with col2:
#     selected_category = st.selectbox("🏷️ Select Category", categories, index=0, key="chart3_category")

# with col3:
#     selected_workspace = st.selectbox("🗂️ Select Workspace ID", workspace_ids, index=0, key="chart3_ws")

# with col1:
#     selected_language = st.selectbox("🗣️ Select Language", languages, index=0, key="chart3_language")



# # --- Time Filter ---
# time_filter = st.radio("⏱ Time Range", ["All Time", "Today", "Yesterday", "Last Week", "Last Month", "Custom Range"],
#                        horizontal=True, key="chart3_time")

# # --- Time Logic ---
# today = dt.date.today()
# start_date, end_date = None, today

# if time_filter == "Today":
#     start_date = today
# elif time_filter == "Yesterday":
#     start_date = today - dt.timedelta(days=1)
#     end_date = start_date
# elif time_filter == "Last Week":
#     start_date = today - dt.timedelta(days=today.weekday() + 7)
#     end_date = start_date + dt.timedelta(days=6)
# elif time_filter == "Last Month":
#     first_day_this_month = today.replace(day=1)
#     start_date = (first_day_this_month - dt.timedelta(days=1)).replace(day=1)
#     end_date = first_day_this_month - dt.timedelta(days=1)
# elif time_filter == "Custom Range":
#     col4, col5 = st.columns(2)
#     with col4:
#         start_date = st.date_input("Start Date", value=today - dt.timedelta(days=30), key="c3_start")
#     with col5:
#         end_date = st.date_input("End Date", value=today, key="c3_end")

# # --- WHERE clause building ---
# where_clauses = []
# params = []

# if selected_user != "All":
#     where_clauses.append("user_first_name = %s")
#     params.append(selected_user)

# if selected_category != "All":
#     where_clauses.append("category = %s")
#     params.append(selected_category)

# if selected_workspace != "All":
#     where_clauses.append("CAST(workspace_id AS TEXT) = %s")
#     params.append(selected_workspace)

# if selected_language != "All":
#     where_clauses.append("language = %s")
#     params.append(selected_language)

# if time_filter != "All Time" and start_date:
#     where_clauses.append("DATE(updated_at) BETWEEN %s AND %s")
#     params.extend([start_date, end_date])

# where_clause = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

# # --- Final SQL Query ---
# query_progress = f"""
# SELECT 
#     DATE(updated_at) AS work_date, 
#     user_first_name,
#     category,
#     COUNT(*) AS annotations_done
# FROM intermediate_table
# {where_clause}
# GROUP BY work_date, user_first_name, category
# ORDER BY work_date ASC;
# """

# # --- Load & Process Data ---
# df_progress = pd.read_sql(query_progress, conn, params=params)
# df_progress["work_date"] = pd.to_datetime(df_progress["work_date"])

# # --- Plotting ---
# fig4 = go.Figure()
# fig4.add_trace(go.Scatter(
#     x=df_progress["work_date"],
#     y=df_progress["annotations_done"],
#     mode="lines+markers",
#     name="Annotations Done",
#     marker=dict(color="#F39C12"),
#     line=dict(width=2)
# ))

# # --- Dynamic Title ---
# title_text = "📅 Daily User Progress"
# if selected_user != "All":
#     title_text += f" for {selected_user}"
# if selected_category != "All":
#     title_text += f" in '{selected_category}'"
# if selected_workspace != "All":
#     title_text += f" (Workspace ID: {selected_workspace})"
# if selected_language != "All":
#     title_text += f" in language '{selected_language}'"
    
# if time_filter != "All Time":
#     title_text += f" ({start_date} to {end_date})"

# fig4.update_layout(
#     title=title_text,
#     xaxis_title="Date",
#     yaxis_title="Annotations Done",
#     plot_bgcolor="rgba(0,0,0,0)",
#     paper_bgcolor="rgba(0,0,0,0)"
# )

# # --- Display Chart ---
# st.plotly_chart(fig4, use_container_width=True, key="chart3_fig")

# # --- Download Button ---
# st.download_button(
#     label="⬇️ Download Chart 3 Data as CSV",
#     data=df_progress.to_csv(index=False),
#     file_name="daily_user_progress_filtered.csv",
#     mime="text/csv",
#     key="chart3_download"
# )

# ------------------- Chart 4: Task Status by Target Language  -------------------
st.markdown("---")
st.markdown("### <a name='chart-4'></a>📈 Chart 4: Task Status by Target Language ", unsafe_allow_html=True)


# Fetch unique languages and task statuses
languages = ["All"] + pd.read_sql("SELECT DISTINCT language FROM intermediate_table WHERE language IS NOT NULL ORDER BY language", conn)["language"].tolist()
statuses = ["annotated", "reviewed", "super_checked", "exported"]

col1, col2 = st.columns(2)

with col1:
    selected_language = st.selectbox("🌐 Select Target Language", languages, key="chart4_lang")

with col2:
    selected_statuses = st.multiselect("✅ Select Task Statuses", statuses, default=statuses, key="chart4_status")

# Time filter
time_filter = st.radio("⏱ Time Range", ["All Time", "Today", "Yesterday", "Last Week", "Last Month", "Custom Range"], horizontal=True, key="chart4_time")

today = dt.date.today()
start_date, end_date = None, today

if time_filter == "Today":
    start_date = today
elif time_filter == "Yesterday":
    start_date = today - dt.timedelta(days=1)
    end_date = start_date
elif time_filter == "Last Week":
    start_date = today - dt.timedelta(days=today.weekday() + 7)
    end_date = start_date + dt.timedelta(days=6)
elif time_filter == "Last Month":
    first_day_this_month = today.replace(day=1)
    start_date = (first_day_this_month - dt.timedelta(days=1)).replace(day=1)
    end_date = first_day_this_month - dt.timedelta(days=1)
elif time_filter == "Custom Range":
    col3, col4 = st.columns(2)
    with col3:
        start_date = st.date_input("Start Date", value=today - dt.timedelta(days=30), key="chart4_start")
    with col4:
        end_date = st.date_input("End Date", value=today, key="chart4_end")

# Build WHERE clause
where_clauses = []
params = []

if selected_language != "All":
    where_clauses.append("language = %s")
    params.append(selected_language)

if selected_statuses and len(selected_statuses) < len(statuses):
    placeholders = ','.join(['%s'] * len(selected_statuses))
    where_clauses.append(f"task_status IN ({placeholders})")
    params.extend(selected_statuses)

if time_filter != "All Time" and start_date:
    where_clauses.append("DATE(updated_at) BETWEEN %s AND %s")
    params.extend([start_date, end_date])

where_clause = "WHERE " + " AND ".join(where_clauses) if where_clauses else ""

# Query with filters
query_lang_status = f"""
SELECT 
    language AS tgt_language,
    COUNT(CASE 
             WHEN task_status IN ('annotated', 'reviewed', 'super_checked', 'exported') 
             THEN 1 
         END) AS annotated_count,
    COUNT(CASE 
             WHEN project_stage = '2' AND task_status IN ('reviewed', 'super_checked', 'exported') THEN 1
             WHEN project_stage > '2' AND task_status IN ('reviewed', 'super_checked') THEN 1
         END) AS reviewed_count,
    COUNT(CASE 
             WHEN project_stage = '3' AND task_status IN ('super_checked', 'exported') THEN 1
             WHEN project_stage != '3' AND task_status = 'super_checked' THEN 1
         END) AS superchecked_count
FROM intermediate_table
{where_clause}
GROUP BY language
ORDER BY reviewed_count DESC;
"""

# Load filtered data
df_lang_status = pd.read_sql(query_lang_status, conn, params=params)

# Plotly Bar Chart (Stacked)
fig6 = go.Figure()
fig6.add_trace(go.Bar(
    x=df_lang_status["tgt_language"],
    y=df_lang_status["annotated_count"],
    name="Annotated",
    marker_color="#3498DB"
))
fig6.add_trace(go.Bar(
    x=df_lang_status["tgt_language"],
    y=df_lang_status["reviewed_count"],
    name="Reviewed",
    marker_color="#2ECC71"
))
fig6.add_trace(go.Bar(
    x=df_lang_status["tgt_language"],
    y=df_lang_status["superchecked_count"],
    name="Superchecked",
    marker_color="#F39C12"
))

# Title logic
title4 = "📊 Task Status by Target Language"
if selected_language != "All":
    title4 += f" - {selected_language}"
if time_filter != "All Time":
    title4 += f" ({start_date} to {end_date})"

fig6.update_layout(
    title=title4,
    barmode="stack",
    xaxis_title="Target Language",
    yaxis_title="Task Count",
    legend_title="Task Status",
    plot_bgcolor="rgba(0,0,0,0)",
    paper_bgcolor="rgba(0,0,0,0)"
)

st.plotly_chart(fig6, use_container_width=True, key="chart4_fig")

# ------------------- Download Chart 4 -------------------
st.download_button(
    label="⬇️ Download Chart 4 Data as CSV",
    data=df_lang_status.to_csv(index=False),
    file_name="task_status_by_target_language_filtered.csv",
    mime="text/csv",
    key="chart4_download"
)

# ------------------- Chart 5: Workspace level Annotation Stats -------------------
import plotly.express as px
st.markdown("---")
st.markdown("### <a name='chart-5'></a>📈 Chart 5:Workspace level Annotation Stats   ", unsafe_allow_html=True)

# ------------------------
# Cached workspace loader
# ------------------------
@st.cache_data
def get_workspaces():
    with conn.cursor() as cur:
        cur.execute("SELECT DISTINCT workspace_id FROM intermediate_table ORDER BY workspace_id;")
        return [row[0] for row in cur.fetchall()]

# -----------------------------------
# Get aggregated annotation stats
# -----------------------------------
def get_annotation_stats(workspace_id):
    query = """
        SELECT 
            project_id,
            COUNT(*) AS total_tasks,
            COUNT(CASE 
                WHEN task_status IN ('annotated', 'reviewed', 'super_checked', 'exported') THEN 1
            END) AS annotated_count,
            COUNT(CASE 
                WHEN project_stage::INTEGER = 2 AND task_status IN ('reviewed', 'super_checked', 'exported') THEN 1
                WHEN project_stage::INTEGER > 2 AND task_status IN ('reviewed', 'super_checked') THEN 1
            END) AS reviewed_count
        FROM intermediate_table
        WHERE workspace_id = %s
        GROUP BY project_id
        ORDER BY project_id;
    """
    return pd.read_sql_query(query, conn, params=(workspace_id,))

# -----------------------------------
# Get drilldown annotation breakdown
# -----------------------------------
def get_annotation_details(project_id):
    # query = """
    #     SELECT
    #         project_id,
    #         annotation_status,
    #         CASE 
    #             WHEN task_status IN ('annotated') THEN 'annotated'
    #             ELSE NULL
    #         END AS annotation_type,
    #         CASE 
    #             WHEN task_status IN ('reviewed') THEN 'reviewed'
    #             ELSE NULL
    #         END AS review_type
    #     FROM intermediate_table
    #     WHERE project_id = %s;
    # """
    query= """SELECT
    project_id,
    annotation_status,
    CASE 
        WHEN annotation_status IN ('labeled', 'unlabeled') THEN 'annotated'
        WHEN annotation_status IN ('accepted_with_minor_changes', 'accepted_with_major_changes', 'unreviewed', 'accepted') THEN NULL
        WHEN task_status = 'annotated' THEN 'annotated'
        ELSE NULL
    END AS annotation_type,
    CASE 
        WHEN annotation_status IN ('accepted_with_minor_changes', 'accepted_with_major_changes', 'unreviewed', 'accepted') THEN 'reviewed'
        WHEN annotation_status IN ('labeled', 'unlabeled') THEN NULL
        WHEN task_status = 'reviewed' THEN 'reviewed'
        ELSE NULL
    END AS review_type
    FROM intermediate_table
    WHERE project_id = %s;
    """
    df = pd.read_sql_query(query, conn, params=(project_id,))
    
    # Create long-format records
    records = []
    for _, row in df.iterrows():
        if row['annotation_type']:
            records.append(('annotated', row['annotation_status']))
        if row['review_type']:
            records.append(('reviewed', row['annotation_status']))

    df_clean = pd.DataFrame(records, columns=['Annotation Type', 'Annotation Status'])

    # Add total for each annotation type
    total_df = (
        df_clean.groupby('Annotation Type')
        .size()
        .reset_index(name='Count')
        .assign(**{'Annotation Status': 'total'})
    )

    combined_df = pd.concat([
        df_clean.value_counts().reset_index(name='Count'),
        total_df
    ], ignore_index=True)

    # Sort 'total' first
    annotation_status_order = ['total'] + sorted(
        [s for s in combined_df['Annotation Status'].unique() if s != 'total']
    )
    combined_df['Annotation Status'] = pd.Categorical(
        combined_df['Annotation Status'],
        categories=annotation_status_order,
        ordered=True
    )

    return combined_df.sort_values(['Annotation Type', 'Annotation Status'])

# ---------------------
# Streamlit UI
# ---------------------

# Workspace selection
workspace_id = st.selectbox("Select a Workspace", get_workspaces())

if workspace_id:
    df = get_annotation_stats(workspace_id)

    if df.empty:
        st.warning("No data found for this workspace.")
    else:
        # -------------------
        # Summary Chart
        # -------------------
        df_melted = df.melt(
            id_vars='project_id',
            value_vars=['total_tasks', 'annotated_count', 'reviewed_count'],
            var_name='Annotation Type',
            value_name='Count'
        )
        df_melted["project_id"] = df_melted["project_id"].astype(str)

        fig = px.bar(
            df_melted,
            x='project_id',
            y='Count',
            color='Annotation Type',
            barmode='group',
            title=f" Annotation Summary for Workspace {workspace_id}",
            text='Count'
        )
        fig.update_layout(
            xaxis_title="Project ID",
            yaxis_title="Count",
            xaxis=dict(type='category'),
            legend_title_text='Annotation Type'
        )
        # fig.update_traces(
        #     textposition='outside',
        #     textangle=0,  # ➤ This makes label horizontal
        #     textfont_size=12
        # )
        fig.update_traces(
            textposition='inside',     # Puts values inside the bars
            textangle=90,             # Rotates text vertically
            textfont_size=12,
            insidetextanchor='middle'  # Keeps text centered inside the bar
        )


        st.plotly_chart(fig, use_container_width=True)
        # Download button for Chart 5
        csv = df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="📥 Download Workspace Annotation Stats as CSV",
            data=csv,
            file_name=f'workspace_{workspace_id}_annotation_stats.csv',
            mime='text/csv'
        )


        # -------------------
        # Drilldown Chart
        # -------------------
        st.markdown("---")
        project_ids = df['project_id'].astype(str).tolist()
        selected_project = st.selectbox(" Select a Project to Drill Down", project_ids)
        if selected_project:
            df_detail = get_annotation_details(int(selected_project))
        
            if df_detail.empty:
                st.info("No annotation data for this project.")
            else:
                # ---------------------------
                # FIX: Prevent bar spacing gaps
                # ---------------------------
        
                annotation_types = ['annotated', 'reviewed']
                annotation_statuses = [
                    'total',
                    'labeled',
                    'unlabeled',
                    'accepted',
                    'accepted_with_major_changes',
                    'accepted_with_minor_changes',
                    'unreviewed'
                ]
        
                # Create all possible combinations of type/status
                full_index = pd.DataFrame(
                    list(product(annotation_types, annotation_statuses)),
                    columns=['Annotation Type', 'Annotation Status']
                )
        
                # Merge to ensure all combinations exist in the dataframe
                df_detail = full_index.merge(df_detail, on=['Annotation Type', 'Annotation Status'], how='left').fillna(0)
                df_detail['Count'] = df_detail['Count'].astype(int)
        
                # Optional: sort bars by logical status order
                df_detail["Annotation Status"] = pd.Categorical(
                    df_detail["Annotation Status"],
                    categories=annotation_statuses,
                    ordered=True
                )
        
                # ---------------------------
                # Create the drilldown chart
                # ---------------------------
                fig_detail = px.bar(
                    df_detail[df_detail["Count"] > 0],  # Only show bars with non-zero values
                    x="Annotation Type",
                    y="Count",
                    color="Annotation Status",
                    barmode="group",
                    text="Count",
                    title=f"Annotation Status Breakdown for Project {selected_project}"
                )
        
                fig_detail.update_layout(
                    xaxis_title="Annotation Type",
                    yaxis_title="Count",
                    xaxis=dict(type='category'),
                    legend_title_text='Annotation Status',
                    bargap=0.15,
                    bargroupgap=0
                )
        
                fig_detail.update_traces(
                    textposition='inside',
                    insidetextanchor='middle',
                    textangle=0
                )
        
                # st.plotly_chart(fig_detail, use_container_width=True)
                st.plotly_chart(fig_detail, use_container_width=True, key=f"plotly_chart_{selected_project}")

        
        # if selected_project:
        #     df_detail = get_annotation_details(int(selected_project))

        #     if df_detail.empty:
        #         st.info("No annotation data for this project.")
        #     else:
        #         # Hide label if count is 0
        #         # Filter out rows with Count == 0 to exclude them from the chart
        #         df_detail = df_detail[df_detail["Count"] > 0]

        #         label_text = df_detail["Count"].apply(lambda x: str(x) if x > 0 else None)
        #         fig_detail = px.bar(
        #             df_detail,
        #             x="Annotation Type",
        #             y="Count",
        #             color="Annotation Status",
        #             barmode="group",
        #             text=df_detail["Count"],
        #             title=f"Annotation Status Breakdown for Project {selected_project}"
        #         )
                
        #         fig_detail.update_layout(
        #             xaxis_title="Annotation Type",
        #             yaxis_title="Count",
        #             xaxis=dict(type='category'),
        #             legend_title_text='Annotation Status',
        #             bargap=0.15,        # Adjust spacing between bars
        #             bargroupgap=0       # No gap between grouped bars
        #         )
                
        #         fig_detail.update_traces(
        #             textposition='inside',
        #             insidetextanchor='middle',
        #             textangle=0
        #         )
# XXXXXXXXXXXXXXXXXXXXXXXXX
                # fig_detail = px.bar(
                #     df_detail,
                #     x="Annotation Type",
                #     y="Count",
                #     color="Annotation Status",
                #     barmode="group",
                #     text=label_text,
                #     title=f" Annotation Status Breakdown for Project {selected_project}"
                # )
                # fig_detail.update_layout(
                #     xaxis_title="Annotation Type",
                #     yaxis_title="Count",
                #     xaxis=dict(type='category'),
                #     legend_title_text='Annotation Status'
                # )
                # fig_detail.update_traces(textposition='outside')
                # st.plotly_chart(fig_detail, use_container_width=True)
# Download button for Drilldown Chart 5
csv_detail = df_detail.to_csv(index=False).encode('utf-8')
st.download_button(
    label="📥 Download Project Annotation Breakdown as CSV",
    data=csv_detail,
    file_name=f'project_{selected_project}_annotation_breakdown.csv',
    mime='text/csv'
)
