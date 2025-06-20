# anudesh_dashboard/charts/excel_export.py
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from io import BytesIO

def generate_language_excel(df2_filtered):
    template_path = "language_report.xlsx"  # Ensure this file exists in working directory
    wb = load_workbook(template_path)
    ws = wb.active

    column_map = {
        ("Creative", "English", "planned"): 2,
        ("Creative", "English", "actual"): 3,
        ("Creative", "Native", "planned"): 4,
        ("Creative", "Native", "actual"): 5,
        ("Creative", "Indic Roman", "planned"): 6,
        ("Creative", "Indic Roman", "actual"): 7,
        ("Cultural", "English", "planned"): 8,
        ("Cultural", "English", "actual"): 9,
        ("Cultural", "Native", "planned"): 10,
        ("Cultural", "Native", "actual"): 11,
        ("Cultural", "Indic Roman", "planned"): 12,
        ("Cultural", "Indic Roman", "actual"): 13,
        ("Education", "English", "planned"): 14,
        ("Education", "English", "actual"): 15,
        ("Education", "Native", "planned"): 16,
        ("Education", "Native", "actual"): 17,
        ("Education", "Indic Roman", "planned"): 18,
        ("Education", "Indic Roman", "actual"): 19,
        ("General", "English", "planned"): 20,
        ("General", "English", "actual"): 21,
        ("General", "Native", "planned"): 22,
        ("General", "Native", "actual"): 23,
        ("General", "Indic Roman", "planned"): 24,
        ("General", "Indic Roman", "actual"): 25,
        ("Safety", "English", "planned"): 26,
        ("Safety", "English", "actual"): 27,
        ("Safety", "Native", "planned"): 28,
        ("Safety", "Native", "actual"): 29,
        ("Safety", "Indic Roman", "planned"): 30,
        ("Safety", "Indic Roman", "actual"): 31,
    }

    languages_list = [
        "Assamese", "Bengali", "Bodo", "Dogri", "Gujarati", "Hindi", "Kannada", "Kashmiri",
        "Maithili", "Malayalam", "Marathi", "Nepali", "Odia", "Punjabi", "Santali", "Tamil",
        "Telugu", "Urdu"
    ]
    language_to_row = {lang: idx + 4 for idx, lang in enumerate(languages_list)}

    for _, row in df2_filtered.iterrows():
        lang = str(row["language"]).strip()
        cat = str(row["category"]).strip()
        stype = str(row["script_type"]).strip()
        planned = pd.to_numeric(row["planned"], errors="coerce")
        actual = pd.to_numeric(row["actual"], errors="coerce")

        if lang in language_to_row:
            row_num = language_to_row[lang]
            for val_type, val in [("planned", planned), ("actual", actual)]:
                col_num = column_map.get((cat, stype, val_type))
                if col_num and pd.notna(val):
                    ws.cell(row=row_num, column=col_num).value = int(val)

    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    for (category, script_type, _), planned_col in column_map.items():
        actual_col = column_map.get((category, script_type, "actual"))
        if not actual_col:
            continue

        for lang in languages_list:
            row = language_to_row[lang]
            planned_cell = ws.cell(row=row, column=planned_col).coordinate
            actual_cell = ws.cell(row=row, column=actual_col).coordinate

            ws.conditional_formatting.add(
                actual_cell,
                CellIsRule(operator='greaterThan', formula=[planned_cell], fill=green_fill)
            )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output